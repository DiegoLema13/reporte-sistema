from flask import Flask, render_template, request
import datetime
import threading
import os
from openpyxl import Workbook, load_workbook
import base64
import json
from zoneinfo import ZoneInfo

import gspread
from google.oauth2.service_account import Credentials

from docxtpl import DocxTemplate
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail,
    Attachment,
    FileContent,
    FileName,
    FileType,
    Disposition,
)

app = Flask(__name__)


# ==============================================
# GENERAR DOCUMENTO WORD DESDE plantilla.docx
# ==============================================
def generar_documento(
    nombre,
    cedula,
    zonal,
    delegada,
    modelo,
    series,
    problema,
    fecha,
    archivo_docx
):

    doc = DocxTemplate("plantilla.docx")

    # Unir todas las series en un solo texto
    series_texto = "\n".join(series)

    #Agregar fecha

    meses = {
    "01": "enero",
    "02": "febrero",
    "03": "marzo",
    "04": "abril",
    "05": "mayo",
    "06": "junio",
    "07": "julio",
    "08": "agosto",
    "09": "septiembre",
    "10": "octubre",
    "11": "noviembre",
    "12": "diciembre",
}

    dia, mes, anio = fecha.split("-")

    fecha_larga = f"{dia} días del mes de {meses[mes]} del {anio}"
    fecha_corta = f"{dia} de {meses[mes]} de {anio}"

    contexto = {
        "nombre": nombre,
        "cedula": cedula,
        "zonal": zonal,
        "delegada": delegada,
        "modelo": modelo,
        "series_texto": series_texto,
        "problema": problema,
        "fecha": fecha,
        "contexto_fecha_larga": fecha_larga,
        "contexto_fecha_corta": fecha_corta,
        
    }
    

    

    print("SERIES TEXTO:")
    print(series_texto)

    doc.render(contexto)
    doc.save(archivo_docx)

# ==============================================
# ENVIAR CORREO CON DOCX ADJUNTO
# ==============================================
def enviar_correo_async(archivo):


    def tarea():
        try:
            message = Mail(
                from_email="ekleain@gmail.com",
                to_emails="ekleain@gmail.com",
                subject="Nuevo reporte técnico",
                html_content="Adjunto reporte técnico generado automáticamente.",
            )

            # Leer el archivo DOCX
            with open(archivo, "rb") as f:
                data = base64.b64encode(f.read()).decode()

            # Adjuntar el archivo Word
            attachment = Attachment(
                FileContent(data),
                FileName(os.path.basename(archivo)),
                FileType(
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                ),
                Disposition("attachment"),
            )

            message.attachment = attachment

            # Enviar usando SendGrid
            sg = SendGridAPIClient(os.environ.get("SENDGRID_API_KEY"))
            response = sg.send(message)

            print("STATUS SENDGRID:", response.status_code)

        except Exception as e:
            print("ERROR SENDGRID:", str(e))

    threading.Thread(target=tarea).start()


    # ==============================================
# REGISTRAR EN EXCEL
# ==============================================
def guardar_excel(
    fecha,
    nombre,
    cedula,
    zonal,
    delegada,
    celular_delegada,
    correo_delegada,
    horario_atencion,
    cantidad_equipos,
    series
):

    archivo_excel = "reportes.xlsx"

    if os.path.exists(archivo_excel):
        wb = load_workbook(archivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

        ws.append([
            "Fecha",
            "Tecnico",
            "Cedula",
            "Zonal",
            "Delegada",
            "Celular",
            "Correo",
            "Horario",
            "Cantidad Equipos",
            "Series"
        ])

    ws.append([
        fecha,
        nombre,
        cedula,
        zonal,
        delegada,
        celular_delegada,
        correo_delegada,
        horario_atencion,
        cantidad_equipos,
        ", ".join(series)
    ])

    wb.save(archivo_excel)

    print("REGISTRO GUARDADO EN EXCEL")

    #==============================================
    #REGISTRAR EN GOOGLE SHEETS
    #==============================================

def guardar_google_sheets(
    fecha,
    nombre,
    cedula,
    zonal,
    delegada,
    cedula_delegada,
    celular_delegada,
    correo_delegada,
    horario_atencion,
    cantidad_equipos,
    series,
    problema
):

    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets"
    ]

    credenciales_json = json.loads(
    os.environ.get("GOOGLE_CREDENTIALS")
)

    creds = Credentials.from_service_account_info(
    credenciales_json,
    scopes=SCOPES
)

    cliente = gspread.authorize(creds)

    hoja = cliente.open_by_key(
        "1kCtmkMrx_7xNFhxACTYfE5K0slokjaL7RJth0hiAQc8"
    ).sheet1

    hoja.append_row([
        fecha,
        nombre,
        cedula,
        zonal,
        delegada,
        cedula_delegada,
        celular_delegada,
        correo_delegada,
        horario_atencion,
        cantidad_equipos,
        ", ".join(series),
        problema
    ])

    print("REGISTRO GUARDADO EN GOOGLE SHEETS")

# ==============================================
# FORMULARIO
# ==============================================
@app.route("/")
def home():
    return render_template("formulario.html")

# ==============================================
# PROCESAR FORMULARIO
# ==============================================
@app.route("/enviar", methods=["POST"])
def enviar():

    # Obtener datos del formulario
    nombre = request.form["nombre"]

    cedula = request.form["cedula"]

    zonal = request.form["zonal"]

    delegada = request.form["delegada"]

    cedula_delegada = request.form["cedula_delegada"]

    celular_delegada = request.form["celular_delegada"]

    correo_delegada = request.form["correo_delegada"]

    horario_atencion = request.form["horario_atencion"]

    problema = request.form["problema"]

    modelo = request.form["modelo"]

    # Obtener múltiples series
    series = request.form.getlist("series[]")

    # Convertir todas las series a mayúsculas
    series = [serie.upper() for serie in series]

    # Cantidad de equipos
    cantidad_equipos = len(series)

    # Mostrar series en consola
    print("SERIES RECIBIDAS:")

    print(series)


    print("CELULAR:", celular_delegada)
    print("CORREO:", correo_delegada)
    print("CEDULA DELEGADA:", cedula_delegada)
    print("HORARIO:", horario_atencion)
    print("CANTIDAD:", cantidad_equipos)

    # Formato de fecha
    fecha = datetime.datetime.now(
    ZoneInfo("America/Guayaquil")).strftime("%d-%m-%Y")

    # Archivo de salidaa
    archivo_docx = "reporte_tecnico.docx"

    

    # Generar documento Word basado en la plantilla
    generar_documento(
        nombre=nombre,
        cedula=cedula,
        zonal=zonal,
        delegada=delegada,
        problema=problema,
        modelo=modelo,
        series=series,
        fecha=fecha,
        archivo_docx=archivo_docx,
    )

    guardar_excel(
    fecha=fecha,
    nombre=nombre,
    cedula=cedula,
    zonal=zonal,
    delegada=delegada,
    celular_delegada=celular_delegada,
    correo_delegada=correo_delegada,
    horario_atencion=horario_atencion,
    cantidad_equipos=cantidad_equipos,
    series=series
)
    guardar_google_sheets(
    fecha=fecha,
    nombre=nombre,
    cedula=cedula,
    zonal=zonal,
    delegada=delegada,
    cedula_delegada=cedula_delegada,
    celular_delegada=celular_delegada,
    correo_delegada=correo_delegada,
    horario_atencion=horario_atencion,
    cantidad_equipos=cantidad_equipos,
    series=series,
    problema=problema
)

    # Enviar el documento por correo
    enviar_correo_async(archivo_docx)

    return render_template("exito.html")


# ==============================================
# EJECUCIÓN LOCAL
# ==============================================
if __name__ == "__main__":
    app.run(debug=True)